import os
from datetime import datetime, timezone
from celery import current_task
from app.tasks.celery_app import celery_app
from app.config import settings


REPORT_STATUS = {}


@celery_app.task(bind=True)
def generate_report(
    self,
    report_type: str,
    start_date: str,
    end_date: str,
    scope: str,
    scope_id: str,
    format: str,
    include: list,
):
    task_id = self.request.id
    REPORT_STATUS[task_id] = {"status": "generating", "progress": 0}

    try:
        from sqlalchemy import create_engine, text

        db_url = settings.database_url.replace("+asyncpg", "")
        engine = create_engine(db_url)

        REPORT_STATUS[task_id] = {"status": "generating", "progress": 30}

        output_dir = os.path.join(settings.backup_path, "reports")
        os.makedirs(output_dir, exist_ok=True)
        filename = f"report_{report_type}_{start_date}_{end_date}_{task_id[:8]}.{format}"
        filepath = os.path.join(output_dir, filename)

        REPORT_STATUS[task_id] = {"status": "generating", "progress": 60}

        if format == "pdf":
            _generate_pdf(filepath, report_type, start_date, end_date, scope)
        elif format == "excel":
            _generate_excel(filepath, report_type, start_date, end_date, scope)
        elif format == "csv":
            _generate_csv(filepath, report_type, start_date, end_date, scope)

        REPORT_STATUS[task_id] = {
            "status": "completed",
            "progress": 100,
            "filepath": filepath,
            "filename": filename,
        }
        return REPORT_STATUS[task_id]

    except Exception as e:
        REPORT_STATUS[task_id] = {"status": "failed", "error": str(e)}
        raise


def _generate_pdf(filepath, report_type, start_date, end_date, scope):
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet

    c = canvas.Canvas(filepath, pagesize=A4)
    width, height = A4

    c.setFont("Helvetica-Bold", 20)
    c.drawString(20 * mm, height - 20 * mm, f"EDU Control Pro - {report_type.title()} Report")
    c.setFont("Helvetica", 12)
    c.drawString(20 * mm, height - 35 * mm, f"Period: {start_date} to {end_date}")
    c.drawString(20 * mm, height - 45 * mm, f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}")
    c.save()


def _generate_excel(filepath, report_type, start_date, end_date, scope):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"EDU Control Pro - {report_type.title()} Report"
    ws["A2"] = f"Period: {start_date} to {end_date}"
    wb.save(filepath)


def _generate_csv(filepath, report_type, start_date, end_date, scope):
    import csv

    with open(filepath, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["EDU Control Pro", report_type, start_date, end_date])
        writer.writerow(["No data available"])


@celery_app.task
def get_report_status(task_id: str):
    return REPORT_STATUS.get(task_id, {"status": "not_found"})
